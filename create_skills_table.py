#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
OpenClaw 技能管理表生成器
"""

import openpyxl
from openpyxl import Workbook
from datetime import datetime

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "OpenClaw 技能管理表"

# 定义表头
headers = [
    "技能名称",
    "分类",
    "技能描述",
    "所需 Key",
    "详细安装和使用方法",
    "注意事项",
    "技能链接",
    "学习状态",
    "安装推荐",
    "学习心得",
    "学习日期"
]

# 写入表头
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# 技能数据
skills_data = [
    {
        "技能名称": "browser-use",
        "分类": "AI 工具",
        "技能描述": "浏览器自动化技能，支持网页抓取、表单填写、UI 测试等功能",
        "所需 Key": "无",
        "详细安装和使用方法": "内置技能，无需安装。使用 browser 工具进行浏览器控制，支持 action=start/snapshot/act/open 等操作",
        "注意事项": "需要配合 snapshot 获取页面元素引用，使用 refs 进行元素定位",
        "技能链接": "https://docs.openclaw.ai",
        "学习状态": "已掌握",
        "安装推荐": "强烈推荐",
        "学习心得": "浏览器自动化核心技能，适合网页数据抓取和 UI 测试",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "agent-browser",
        "分类": "AI 工具",
        "技能描述": "基于浏览器的智能体技能，可在浏览器环境中运行 AI 任务",
        "所需 Key": "无",
        "详细安装和使用方法": "通过 browser 工具实现，支持在隔离浏览器中运行自动化任务",
        "注意事项": "需要合理设置超时和等待时间，避免页面加载问题",
        "技能链接": "https://docs.openclaw.ai",
        "学习状态": "学习中",
        "安装推荐": "推荐",
        "学习心得": "适合需要登录状态或复杂交互的网页任务",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "active-maintenance",
        "分类": "其他",
        "技能描述": "主动维护机制，包括心跳检查、定期任务执行等",
        "所需 Key": "无",
        "详细安装和使用方法": "通过 HEARTBEAT.md 配置定期任务，使用 cron 工具管理定时任务",
        "注意事项": "避免过于频繁的检查，注意 API 调用频率限制",
        "技能链接": "https://docs.openclaw.ai",
        "学习状态": "已掌握",
        "安装推荐": "强烈推荐",
        "学习心得": "保持助手主动性的关键机制，建议配置邮件、日历等定期检查",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "agent-commons",
        "分类": "开发工具",
        "技能描述": "智能体通用组件库，提供常用功能和工具",
        "所需 Key": "无",
        "详细安装和使用方法": "作为基础依赖包使用，提供会话管理、工具调用等通用功能",
        "注意事项": "了解各组件的依赖关系，避免版本冲突",
        "技能链接": "https://github.com/openclaw/openclaw",
        "学习状态": "学习中",
        "安装推荐": "推荐",
        "学习心得": "理解底层架构有助于更好地使用 OpenClaw",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "feishu-bitable-creator",
        "分类": "AI 工具",
        "技能描述": "飞书多维表格创建和管理技能",
        "所需 Key": "FEISHU_APP_ID, FEISHU_APP_SECRET",
        "详细安装和使用方法": "1. 在飞书开放平台创建应用 2. 获取 App ID 和 Secret 3. 配置到环境变量 4. 使用 feishu-drive 或 feishu-doc 技能",
        "注意事项": "需要飞书企业管理员授权，注意权限范围",
        "技能链接": "https://open.feishu.cn",
        "学习状态": "学习中",
        "安装推荐": "推荐",
        "学习心得": "飞书生态重要组成部分，适合团队协作场景",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "ai-news-oracle",
        "分类": "搜索工具",
        "技能描述": "AI 新闻聚合和摘要技能，提供每日 AI 资讯",
        "所需 Key": "无（或使用 digest-to-email）",
        "详细安装和使用方法": "使用 digest-to-email 技能生成每日摘要，或通过 web_search 获取最新新闻",
        "注意事项": "注意新闻源的可靠性和时效性",
        "技能链接": "https://clawhub.com",
        "学习状态": "已掌握",
        "安装推荐": "推荐",
        "学习心得": "保持技术敏感度的好工具，建议配置每日定时推送",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "bug-audit",
        "分类": "开发工具",
        "技能描述": "代码 Bug 审计和安全检查技能",
        "所需 Key": "无",
        "详细安装和使用方法": "使用 coding-agent 技能配合代码分析工具，或手动审查代码",
        "注意事项": "结合人工审查，自动化检查可能有误报",
        "技能链接": "https://github.com/openclaw/openclaw",
        "学习状态": "学习中",
        "安装推荐": "推荐",
        "学习心得": "代码质量保障的重要环节",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "nodejs-project-arch",
        "分类": "开发工具",
        "技能描述": "Node.js 项目架构模板和最佳实践",
        "所需 Key": "无",
        "详细安装和使用方法": "参考 OpenClaw 项目结构，使用标准 Node.js 项目模板",
        "注意事项": "遵循项目规范，保持代码结构清晰",
        "技能链接": "https://github.com/openclaw/openclaw",
        "学习状态": "学习中",
        "安装推荐": "可选",
        "学习心得": "良好的项目结构有助于长期维护",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "coze-web-search",
        "分类": "搜索工具",
        "技能描述": "Coze 平台网页搜索技能",
        "所需 Key": "COZE_API_KEY",
        "详细安装和使用方法": "1. 在 Coze 平台创建 Bot 2. 配置搜索插件 3. 获取 API Key 4. 配置到环境变量",
        "注意事项": "Coze API 可能有调用限制",
        "技能链接": "https://www.coze.cn",
        "学习状态": "未开始",
        "安装推荐": "可选",
        "学习心得": "Coze 生态的搜索能力补充",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "coze-web-fetch",
        "分类": "搜索工具",
        "技能描述": "Coze 平台网页内容抓取技能",
        "所需 Key": "COZE_API_KEY",
        "详细安装和使用方法": "类似 coze-web-search，配置 Coze Bot 和 API Key",
        "注意事项": "注意目标网站的 robots.txt 和使用条款",
        "技能链接": "https://www.coze.cn",
        "学习状态": "未开始",
        "安装推荐": "可选",
        "学习心得": "网页内容抓取的另一种选择",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "news-aggregator-skill",
        "分类": "搜索工具",
        "技能描述": "新闻聚合技能，整合多源新闻内容",
        "所需 Key": "无",
        "详细安装和使用方法": "使用 web_search 或 multi-search-engine 技能聚合新闻源",
        "注意事项": "注意新闻源的多样性和平衡性",
        "技能链接": "https://clawhub.com",
        "学习状态": "学习中",
        "安装推荐": "推荐",
        "学习心得": "信息获取的重要渠道",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "feishu-sheets",
        "分类": "AI 工具",
        "技能描述": "飞书表格操作技能，支持读写和管理",
        "所需 Key": "FEISHU_APP_ID, FEISHU_APP_SECRET",
        "详细安装和使用方法": "1. 创建飞书应用 2. 配置表格权限 3. 获取凭证 4. 使用 feishu-doc/feishu-drive 技能",
        "注意事项": "需要适当的表格访问权限",
        "技能链接": "https://open.feishu.cn",
        "学习状态": "学习中",
        "安装推荐": "推荐",
        "学习心得": "飞书协作的核心功能之一",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "frontend-design-skill",
        "分类": "开发工具",
        "技能描述": "前端设计技能，支持 UI/UX 设计和实现",
        "所需 Key": "无",
        "详细安装和使用方法": "结合 browser 工具和前端框架，使用 coding-agent 进行开发",
        "注意事项": "关注响应式设计和用户体验",
        "技能链接": "https://clawhub.com",
        "学习状态": "学习中",
        "安装推荐": "可选",
        "学习心得": "前端开发的基础能力",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "agent-team-orchestration",
        "分类": "AI 工具",
        "技能描述": "多智能体协同编排技能，管理多个子代理",
        "所需 Key": "无",
        "详细安装和使用方法": "使用 sessions_spawn 创建子代理，通过 subagents 工具进行管理和协调",
        "注意事项": "合理分配任务，避免资源浪费",
        "技能链接": "https://docs.openclaw.ai",
        "学习状态": "已掌握",
        "安装推荐": "强烈推荐",
        "学习心得": "复杂任务分解和并行处理的关键",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "arxiv-search-collector",
        "分类": "搜索工具",
        "技能描述": "arXiv 学术论文搜索和收集技能",
        "所需 Key": "无",
        "详细安装和使用方法": "使用 web_search 搜索 arxiv.org，或使用 baidu-scholar-search-skill 获取学术文献",
        "注意事项": "注意论文的引用格式和版权",
        "技能链接": "https://arxiv.org",
        "学习状态": "学习中",
        "安装推荐": "推荐",
        "学习心得": "学术研究的重要工具",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "三层记忆架构",
        "分类": "其他",
        "技能描述": "OpenClaw 核心记忆系统：短期会话记忆、中期日记记忆、长期 MEMORY.md",
        "所需 Key": "无",
        "详细安装和使用方法": "1. 会话记忆自动管理 2. 创建 memory/YYYY-MM-DD.md 记录日常 3. 维护 MEMORY.md 存储长期知识",
        "注意事项": "定期回顾和整理记忆文件，避免信息过载",
        "技能链接": "https://docs.openclaw.ai",
        "学习状态": "已掌握",
        "安装推荐": "强烈推荐",
        "学习心得": "理解记忆架构是掌握 OpenClaw 的关键，每次会话前读取相关文件",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "OpenClaw 心跳机制",
        "分类": "其他",
        "技能描述": "主动检查和提醒机制，通过 HEARTBEAT.md 配置定期任务",
        "所需 Key": "无",
        "详细安装和使用方法": "1. 编辑 HEARTBEAT.md 添加检查任务 2. 配置 cron 定时任务 3. 实现邮件、日历等检查",
        "注意事项": "避免过于频繁的检查，注意 API 限额",
        "技能链接": "https://docs.openclaw.ai",
        "学习状态": "已掌握",
        "安装推荐": "强烈推荐",
        "学习心得": "保持助手主动性的核心机制，建议配置 2-4 次/天的检查频率",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "Skill 最小可用面",
        "分类": "开发工具",
        "技能描述": "技能开发的最小可行产品概念，快速创建可用技能",
        "所需 Key": "无",
        "详细安装和使用方法": "1. 参考 skill-creator 技能 2. 创建 SKILL.md 定义技能 3. 实现核心功能 4. 测试和迭代",
        "注意事项": "先实现核心功能，再逐步完善",
        "技能链接": "https://docs.openclaw.ai",
        "学习状态": "学习中",
        "安装推荐": "推荐",
        "学习心得": "快速原型开发的好方法",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "多技能协同编排",
        "分类": "AI 工具",
        "技能描述": "协调多个技能完成复杂任务的能力",
        "所需 Key": "无",
        "详细安装和使用方法": "1. 理解各技能功能 2. 设计任务流程 3. 按顺序调用技能 4. 处理技能间数据传递",
        "注意事项": "注意技能依赖和执行顺序",
        "技能链接": "https://docs.openclaw.ai",
        "学习状态": "已掌握",
        "安装推荐": "强烈推荐",
        "学习心得": "OpenClaw 的核心优势，通过技能组合实现复杂工作流",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "学术雷达 V3.0",
        "分类": "搜索工具",
        "技能描述": "学术文献追踪和推荐系统，整合多源学术资源",
        "所需 Key": "无（或使用 baidu-scholar-search-skill）",
        "详细安装和使用方法": "使用 baidu-scholar-search-skill 搜索中英文学术文献，结合 web_search 获取最新研究",
        "注意事项": "注意学术资源的可访问性和版权",
        "技能链接": "https://clawhub.com",
        "学习状态": "学习中",
        "安装推荐": "推荐",
        "学习心得": "学术研究的好帮手，建议定期使用",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    },
    {
        "技能名称": "飞书文档创建最佳实践",
        "分类": "AI 工具",
        "技能描述": "飞书文档创建和管理的最佳实践指南",
        "所需 Key": "FEISHU_APP_ID, FEISHU_APP_SECRET",
        "详细安装和使用方法": "1. 配置飞书应用凭证 2. 使用 feishu-doc 技能创建文档 3. 使用 feishu-perm 管理权限 4. 使用 feishu-wiki 组织知识库",
        "注意事项": "合理设置文档权限，注意内容安全",
        "技能链接": "https://open.feishu.cn",
        "学习状态": "学习中",
        "安装推荐": "推荐",
        "学习心得": "飞书协作的基础，建议配合 feishu-drive 使用",
        "学习日期": datetime.now().strftime("%Y-%m-%d")
    }
]

# 写入数据
for row_idx, skill in enumerate(skills_data, 2):
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=row_idx, column=col_idx, value=skill[header])

# 设置列宽
column_widths = [15, 12, 30, 20, 40, 25, 25, 12, 12, 30, 15]
for col, width in enumerate(column_widths, 1):
    ws.column_dimensions[chr(64 + col)].width = width

# 设置表头样式
from openpyxl.styles import Font, PatternFill, Alignment

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center")

for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# 设置数据行样式
data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
for row in range(2, len(skills_data) + 2):
    for col in range(1, len(headers) + 1):
        ws.cell(row=row, column=col).alignment = data_alignment

# 保存文件
output_path = r"C:\Users\Administrator\.openclaw\workspace\OpenClaw 技能管理表.xlsx"
wb.save(output_path)

print(f"[OK] 技能管理表已生成：{output_path}")
print(f"[INFO] 共包含 {len(skills_data)} 个技能")
