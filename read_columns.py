#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""使用 openpyxl 读取参考 Excel 文件的列名"""

from openpyxl import load_workbook

# 读取参考文件
ref_path = r"C:\Users\Administrator\.openclaw\media\outbound\7edfac85-80ef-49fd-9750-05e75f1c3c3e.xlsx"

wb = load_workbook(ref_path, data_only=True)
ws = wb.active

print("=" * 80)
print("列名 (Columns):")
print("=" * 80)
columns = []
for i in range(1, 12):
    col_name = ws.cell(row=1, column=i).value
    columns.append(col_name)
    print(f"{i}. {col_name}")

print("\n" + "=" * 80)
print("第一行数据示例:")
print("=" * 80)
for i, col in enumerate(columns, 1):
    val = ws.cell(row=2, column=i).value
    print(f"{col}: {val}")

print("\n" + "=" * 80)
print("总行数:", ws.max_row)
print("=" * 80)
