#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
OpenClaw 技能管理表 - 更新技能链接
"""

import openpyxl
from datetime import datetime

# 加载现有工作簿
input_path = r"C:\Users\Administrator\.openclaw\workspace\OpenClaw 技能管理表.xlsx"
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# 技能链接数据（基于 skills.sh、GitHub、coze 等平台）
skills_links = {
    "browser-use": "https://github.com/browser-use/browser-use\nhttps://skills.sh/browser-use/browser-use/browser-use",
    "agent-browser": "https://github.com/vercel-labs/agent-browser\nhttps://skills.sh/vercel-labs/agent-browser/agent-browser",
    "active-maintenance": "https://docs.openclaw.ai\nhttps://github.com/openclaw/openclaw",
    "agent-commons": "https://github.com/openclaw/openclaw\nhttps://docs.openclaw.ai",
    "feishu-bitable-creator": "https://open.feishu.cn/document/home/index\nhttps://github.com/search?q=feishu+bitable+skill",
    "ai-news-oracle": "https://clawhub.ai\nhttps://github.com/search?q=ai+news+oracle+skill",
    "bug-audit": "https://github.com/search?q=bug+audit+skill\nhttps://skills.sh/squirrelscan/skills/audit-website",
    "nodejs-project-arch": "https://github.com/search?q=nodejs+project+template\nhttps://github.com/openclaw/openclaw",
    "coze-web-search": "https://www.coze.cn\nhttps://instreet.coze.site/",
    "coze-web-fetch": "https://www.coze.cn\nhttps://instreet.coze.site/",
    "news-aggregator-skill": "https://clawhub.ai\nhttps://github.com/search?q=news+aggregator+skill",
    "feishu-sheets": "https://open.feishu.cn/document/home/index\nhttps://github.com/search?q=feishu+sheets+skill",
    "frontend-design-skill": "https://github.com/anthropics/skills\nhttps://skills.sh/anthropics/skills/frontend-design",
    "agent-team-orchestration": "https://docs.openclaw.ai\nhttps://github.com/openclaw/openclaw",
    "arxiv-search-collector": "https://arxiv.org\nhttps://github.com/search?q=arxiv+search+skill",
    "三层记忆架构": "https://docs.openclaw.ai\nhttps://github.com/openclaw/openclaw",
    "OpenClaw 心跳机制": "https://docs.openclaw.ai\nhttps://github.com/openclaw/openclaw",
    "Skill 最小可用面": "https://github.com/anthropics/skills\nhttps://skills.sh/anthropics/skills/skill-creator",
    "多技能协同编排": "https://docs.openclaw.ai\nhttps://github.com/openclaw/openclaw",
    "学术雷达 V3.0": "https://clawhub.ai\nhttps://github.com/search?q=academic+radar+skill",
    "飞书文档创建最佳实践": "https://open.feishu.cn/document/home/index\nhttps://github.com/search?q=feishu+doc+skill"
}

# 找到技能名称列和链接列的索引
headers = [cell.value for cell in ws[1]]
skill_name_idx = headers.index("技能名称") + 1
skill_link_idx = headers.index("技能链接") + 1

# 更新链接
updated_count = 0
for row in range(2, ws.max_row + 1):
    skill_name = ws.cell(row=row, column=skill_name_idx).value
    if skill_name in skills_links:
        ws.cell(row=row, column=skill_link_idx).value = skills_links[skill_name]
        updated_count += 1
        print(f"[OK] 已更新：{skill_name}")

# 设置列宽
ws.column_dimensions['G'].width = 50  # 技能链接列

# 保存文件
wb.save(input_path)

print(f"\n[OK] 技能链接更新完成！")
print(f"[INFO] 共更新 {updated_count} 个技能的链接")
print(f"[INFO] 文件位置：{input_path}")
