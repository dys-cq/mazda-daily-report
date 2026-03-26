#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
根据参考文件样式创建已安装技能列表
列名：技能名称 | 分类 | 技能描述 | 所需 Key | 详细安装和使用方法 | 注意事项 | 学习状态 | 安装推荐 | 学习心得 | 学习日期 | 技能链接 URL
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from datetime import date

# 参考已有数据推断列名（从参考文件）
# 1.技能名称 2.分类 3.技能描述 4.所需 Key 5.详细安装和使用方法 6.注意事项 7.学习状态 8.安装推荐 9.学习心得 10.学习日期 11.技能链接 URL

# 技能数据（37 个已安装技能）
skills_data = [
    # baoyu 系列 (16 个) - 来自 github.com/ideacco/baoyu-skills-openclaw
    ["baoyu-article-illustrator", "AI 创作", "分析文章结构，识别需要配图的位置，生成类型×风格二维插图", "否", "使用该技能，无需安装。使用 baoyu 工具进行文章插图生成，支持类型×风格二维方式", "需要配合 baoyu-image-gen 等后端技能使用", "未开始", "推荐", "", "", "https://github.com/ideacco/baoyu-skills-openclaw"],
    ["baoyu-comic", "AI 创作", "知识漫画创作者，支持多种艺术风格和语气，创建原创教育漫画", "否", "使用该技能，无需安装。通过自然语言描述漫画主题和风格，自动生成多格漫画", "生成时间较长，需要耐心等待", "未开始", "推荐", "", "", "https://github.com/ideacco/baoyu-skills-openclaw"],
    ["baoyu-compress-image", "工具", "压缩图片为 WebP（默认）或 PNG 格式，自动选择工具", "否", "使用该技能，无需安装。支持指定质量参数和输出格式", "macOS 建议安装 webp (brew install webp) 以获得更好性能", "未开始", "强烈推荐", "图片压缩必备工具，适合批量处理", date.today().strftime("%Y-%m-%d"), "https://github.com/ideacco/baoyu-skills-openclaw"],
    ["baoyu-cover-image", "AI 创作", "生成文章封面图，支持 5 维度定制（类型、调色板、渲染、文字、情绪）", "否", "使用该技能，无需安装。支持电影比例 (2.35:1)、宽屏 (16:9)、方形 (1:1)", "需要配合图像生成后端", "未开始", "推荐", "", "", "https://github.com/ideacco/baoyu-skills-openclaw"],
    ["baoyu-danger-gemini-web", "AI 创作", "通过逆向 Gemini Web API 生成图像和文本", "否", "使用该技能，无需安装。支持文本生成、图像生成、参考图输入和多轮对话", "逆向 API，可能存在稳定性问题", "未开始", "推荐", "", "", "https://github.com/ideacco/baoyu-skills-openclaw"],
    ["baoyu-danger-x-to-markdown", "工具", "将 X (Twitter) 推文和文章转换为 Markdown 格式", "否", "使用该技能，无需安装。提供 x.com/twitter.com URL 即可转换", "需要用户同意使用逆向 API", "未开始", "推荐", "", "", "https://github.com/ideacco/baoyu-skills-openclaw"],
    ["baoyu-format-markdown", "工具", "格式化纯文本或 Markdown 文件，添加 frontmatter、标题、摘要等", "否", "使用该技能，无需安装。输出到 {filename}-formatted.md", "首次使用需要安装依赖", "未开始", "强烈推荐", "文章格式化必备", date.today().strftime("%Y-%m-%d"), "https://github.com/ideacco/baoyu-skills-openclaw"],
    ["baoyu-image-gen", "AI 创作", "AI 图像生成，支持 OpenAI、Google、DashScope 和 Replicate API", "OPENAI_API_KEY/GOOGLE_API_KEY/DASHSCOPE_API_KEY/REPLICATE_API_TOKEN", "使用该技能，无需安装。支持文本到图像、参考图、多种宽高比", "需要至少一个 API Key；批量生成建议使用 batch 模式", "未开始", "强烈推荐", "图像生成核心技能", date.today().strftime("%Y-%m-%d"), "https://github.com/ideacco/baoyu-skills-openclaw"],
    ["baoyu-infographic", "AI 创作", "生成专业信息图，21 种布局类型和 20 种视觉风格", "否", "使用该技能，无需安装。分析内容后推荐布局×风格组合", "内容较多时生成时间较长", "未开始", "推荐", "", "", "https://github.com/ideacco/baoyu-skills-openclaw"],
    ["baoyu-markdown-to-html", "工具", "将 Markdown 转换为带样式的 HTML，支持微信兼容主题", "否", "使用该技能，无需安装。支持代码高亮、数学公式、PlantUML、脚注等", "首次使用需要安装依赖", "未开始", "强烈推荐", "公众号文章发布必备", date.today().strftime("%Y-%m-%d"), "https://github.com/ideacco/baoyu-skills-openclaw"],
    ["baoyu-post-to-wechat", "社交媒体", "微信公众号文章一键发布工具，支持专业排版、AI 封面图生成", "否", "使用该技能，无需安装。通过 Chrome CDP 自动发布到草稿箱", "需要登录微信公众号后台；首次使用需要配置", "未开始", "强烈推荐", "公众号发布自动化核心", date.today().strftime("%Y-%m-%d"), "https://github.com/ideacco/baoyu-skills-openclaw"],
    ["baoyu-post-to-wechat-backup-20260315_014801", "社交媒体", "微信公众号发布工具备份版本", "否", "备份版本，不建议使用。使用主版本 baoyu-post-to-wechat", "备份版本，已过时", "未开始", "不推荐", "使用主版本即可", "", "https://github.com/ideacco/baoyu-skills-openclaw"],
    ["baoyu-post-to-weibo", "社交媒体", "发布内容到微博，支持普通帖子和头条文章", "否", "使用该技能，无需安装。通过 Chrome CDP 自动发布", "需要登录微博账号", "未开始", "推荐", "", "", "https://github.com/ideacco/baoyu-skills-openclaw"],
    ["baoyu-post-to-x", "社交媒体", "发布内容和文章到 X (Twitter)，支持普通帖子和 X Articles", "否", "使用该技能，无需安装。通过 Chrome CDP 绕过反自动化", "需要登录 X 账号；X Articles 需要 Premium", "未开始", "推荐", "", "", "https://github.com/ideacco/baoyu-skills-openclaw"],
    ["baoyu-slide-deck", "AI 创作", "从内容生成专业幻灯片图片", "否", "使用该技能，无需安装。创建大纲和风格说明，然后生成单张幻灯片图片", "生成时间较长", "未开始", "推荐", "", "", "https://github.com/ideacco/baoyu-skills-openclaw"],
    ["baoyu-url-to-markdown", "工具", "使用 Chrome CDP 获取任何 URL 并转换为 Markdown", "否", "使用该技能，无需安装。支持自动捕获和等待用户信号两种模式", "需要本地 Chrome；某些页面需要登录", "未开始", "强烈推荐", "网页保存必备", date.today().strftime("%Y-%m-%d"), "https://github.com/ideacco/baoyu-skills-openclaw"],
    ["baoyu-xhs-images", "AI 创作", "生成小红书信息图系列，11 种视觉风格和 8 种布局", "否", "使用该技能，无需安装。将内容分解为 1-10 张卡通风格图片", "针对小红书平台优化", "未开始", "推荐", "", "", "https://github.com/ideacco/baoyu-skills-openclaw"],
    
    # 百度系列
    ["baidu-baike-data", "搜索查询", "百度百科查询工具，查询权威百科名词解释", "BAIDU_API_KEY", "1. 安装：clawhub install baidu-baike-data\n2. 直接查询：python3 scripts/baidu_baike.py --search_type=lemmaTitle --search_key=\"关键词\"\n3. 查询列表：python3 scripts/baidu_baike.py --search_type=lemmaList --search_key=\"关键词\"\n4. 指定 ID: python3 scripts/baidu_baike.py --search_type=lemmaId --search_key=\"entry_id\"", "查询术语需要先查询 lemmaList 获取列表，再查询 lemmaId", "未开始", "推荐", "适合查询固定术语", "", "https://clawhub.com/skills/baidu-baike-data"],
    ["baidu-scholar-search-skill", "搜索查询", "百度学术搜索 - 搜索中英文学术文献（期刊、会议、论文等）", "BAIDU_API_KEY", "1. 安装：clawhub install baidu-scholar-search-skill\n2. 基本搜索：bash baidu_scholar_search.sh \"关键词\"\n3. 带摘要：bash baidu_scholar_search.sh \"关键词\" 0 true\n4. 分页：bash baidu_scholar_search.sh \"关键词\" 1", "带摘要会增加响应时间；页码从 0 开始", "未开始", "推荐", "", "", "https://clawhub.com/skills/baidu-scholar-search-skill"],
    ["baidu-search", "搜索查询", "使用百度 AI 搜索引擎进行网络搜索", "BAIDU_API_KEY", "1. 安装：clawhub install baidu-search\n2. 基本搜索：python3 scripts/search.py '{\"query\":\"关键词\"}'\n3. 时间筛选：python3 scripts/search.py '{\"query\":\"新闻\",\"freshness\":\"pd\"}'\n4. 更多结果：python3 scripts/search.py '{\"query\":\"新闻\",\"count\":20}'", "freshness 支持 pd(24h)、pw(7 天)、pm(31 天)、py(365 天)；超过时间范围无效", "未开始", "强烈推荐", "百度生态内容搜索首选", date.today().strftime("%Y-%m-%d"), "https://clawhub.com/skills/baidu-search"],
    
    # 飞书系列 (4 个) - 来自 github.com/autogame-17/feishu-skills
    ["feishu-doc", "办公协作", "飞书文档读写操作", "否", "使用该技能，无需安装。支持 read/write/append/create/list_blocks/get_block/update_block/delete_block/create_table/write_table_cells 等 action", "Markdown 表格不支持；结构化内容需要使用 list_blocks", "未开始", "强烈推荐", "飞书文档操作核心技能", date.today().strftime("%Y-%m-%d"), "https://github.com/autogame-17/feishu-skills"],
    ["feishu-drive", "办公协作", "飞书云存储文件管理", "否", "使用该技能，无需安装。支持 list/info/create_folder/move/delete 等 action", "文件必须在根目录或使用 folder_token 指定文件夹", "未开始", "推荐", "", "", "https://github.com/autogame-17/feishu-skills"],
    ["feishu-perm", "办公协作", "飞书文档和文件权限管理", "否", "使用该技能，无需安装。支持添加/移除协作者、设置权限等级", "需要文档 token 和用户 open_id", "未开始", "推荐", "", "", "https://github.com/autogame-17/feishu-skills"],
    ["feishu-wiki", "办公协作", "飞书知识库导航", "否", "使用该技能，无需安装。通过 feishu_doc 工具读写知识库页面内容", "依赖 feishu-doc 工具", "未开始", "推荐", "", "", "https://github.com/autogame-17/feishu-skills"],
    
    # 其他技能
    ["anything-to-notebooklm", "AI 创作", "多源内容智能处理器：支持微信公众号、网页、YouTube、PDF、Markdown 等，自动上传到 NotebookLM 并生成播客/PPT/思维导图等多种格式", "否", "使用该技能，无需安装。支持多种内容源自动处理和上传", "需要 NotebookLM 账号", "未开始", "推荐", "", "", ""],
    ["baoyu-translate", "工具", "文章和文档翻译，支持快速/普通/精翻三种模式", "否", "使用该技能，无需安装。支持自定义术语表和 EXTEND.md 配置", "精翻模式需要更多时间", "未开始", "强烈推荐", "多语言内容处理必备", date.today().strftime("%Y-%m-%d"), ""],
    ["baoyu-yunwu", "AI 创作", "通过 Yunwu AI API 调用 Gemini Imagen 模型生成图片", "YUNWU_API_KEY", "使用该技能，无需安装。支持 16:9、9:16、1:1、900x383 等比例，1K/2K 质量", "需要 Yunwu API Key", "未开始", "推荐", "", "", ""],
    ["digest-to-email", "自动化", "生成每日 AI/技术摘要并通过 SMTP 发送邮件", "SMTP 配置", "使用该技能，无需安装。可选加载 ~/.hn-daily-digest/email_config.json 中的 SMTP 配置", "需要配置 SMTP 凭据", "未开始", "推荐", "", "", ""],
    ["html-to-media-cover", "工具", "HTML 转长图并适配自媒体封面尺寸", "否", "使用该技能，无需安装。适配小红书/公众号等平台封面尺寸", "需要本地 Chrome", "未开始", "推荐", "", "", ""],
    ["info-designer-infographic", "工具", "高密度信息设计与视觉生成指导", "否", "使用该技能，无需安装。提供信息图设计指导而非实际生成", "设计指导类技能", "未开始", "推荐", "", "", ""],
    ["markdown-mailer", "自动化", "通过 SMTP 发送 Markdown 或 HTML 内容邮件", "SMTP 配置", "使用该技能，无需安装。可选加载 ~/.hn-daily-digest/email_config.json 中的 SMTP 配置", "需要配置 SMTP 凭据", "未开始", "推荐", "", "", ""],
    ["mazda-daily-report", "自动化", "生成马自达售后日报（经营 + 客诉/线索+CSI）", "否", "使用该技能，无需安装。读取 E:\\每日分析数据源 下 Excel 源表，按固定 5 家店统计", "特定业务场景技能", "未开始", "不推荐", "仅限马自达业务使用", "", ""],
    ["miaoda-app-builder", "开发工具", "在百度秒哒平台创建、修改、生成和部署网站和 Web 应用", "MIAODA_API_KEY", "1. 安装：clawhub install miaoda-app-builder\n2. 配置 Key: export MIAODA_API_KEY=\"sk_xxxxx\"\n3. 创建应用：python scripts/miaoda_api.py chat --text \"创建一个 XX 应用\"\n4. 发布：python scripts/miaoda_api.py publish --app-id <id> --wait", "需要秒哒平台 API Key；生成后使用 chat 修改，不要重复 generate-app", "未开始", "强烈推荐", "", "", "https://clawhub.com/skills/miaoda-app-builder"],
    ["multi-search-engine", "搜索查询", "多搜索引擎集成，17 个引擎（8 个国内 +9 个全球）", "否", "使用该技能，无需安装。支持高级搜索运算符、时间筛选、站内搜索、隐私引擎和 WolframAlpha 查询", "无需 API Key", "未开始", "强烈推荐", "无 API 依赖的搜索方案", date.today().strftime("%Y-%m-%d"), "https://clawhub.com/skills/multi-search-engine"],
    ["OpenClaw-DeepReeder", "工具", "OpenClaw 代理的默认网页内容网关，读取 X、Reddit、YouTube 等", "否", "自动运行，无需安装。Paste URL 到对话中自动获取、解析和保存高质量 Markdown", "零配置、零 API Key", "未开始", "强烈推荐", "网页内容读取核心", date.today().strftime("%Y-%m-%d"), "https://github.com/astonysh/OpenClaw-DeepReeder"],
    ["skill-creator", "工具", "创建或更新 AgentSkills", "否", "使用该技能，无需安装。用于设计、构建和打包技能", "技能开发专用", "未开始", "推荐", "", "", "https://clawhub.com/skills/skill-creator"],
    ["weather", "工具", "通过 wttr.in 或 Open-Meteo 获取当前天气和预报", "否", "使用该技能，无需安装。支持任何地点的当前天气和预报查询", "不支持历史天气数据、严重天气警报", "未开始", "强烈推荐", "日常天气查询", date.today().strftime("%Y-%m-%d"), "https://clawhub.com/skills/weather"],
    
    # 其他 GitHub 技能
    ["Agent-Reach", "搜索查询", "给 AI Agent 装上互联网能力：支持 13+ 平台（Twitter、Reddit、YouTube、GitHub、B 站、小红书、抖音等）的读取和搜索", "否", "1. 一键安装：告诉 Agent「帮我安装 Agent Reach：https://raw.githubusercontent.com/Panniantong/agent-reach/main/docs/install.md」\n2. 更新：「帮我更新 Agent Reach：https://raw.githubusercontent.com/Panniantong/agent-reach/main/docs/update.md」\n3. 诊断：agent-reach doctor\n4. 卸载：agent-reach uninstall", "Cookie 登录的平台存在封号风险，建议使用小号；OpenClaw 用户需要先开启 exec 权限", "未开始", "强烈推荐", "9.4k stars 的热门项目，AI Agent 互联网访问必备工具", date.today().strftime("%Y-%m-%d"), "https://github.com/Panniantong/Agent-Reach"],
]

# 列名（与参考文件一致）
columns = [
    "技能名称",
    "分类",
    "技能描述",
    "所需 Key",
    "详细安装和使用方法",
    "注意事项",
    "学习状态",
    "安装推荐",
    "学习心得",
    "学习日期",
    "技能链接 URL"
]

# 创建 DataFrame
df = pd.DataFrame(skills_data, columns=columns)

# 导出到 Excel
output_path = r"C:\Users\Administrator\.openclaw\workspace\OpenClaw 技能管理表_完整版_v3.xlsx"

# 使用 ExcelWriter 创建文件
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name="已安装技能")
    
    # 获取 workbook 和 worksheet
    wb = writer.book
    ws = wb.active
    
    # 设置列宽（根据参考文件）
    column_widths = [25, 12, 50, 15, 70, 40, 12, 12, 50, 15, 70]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 设置数据行样式
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(columns)):
        for cell in row:
            cell.alignment = data_alignment
            cell.border = border
    
    # 冻结首行
    ws.freeze_panes = 'A2'

print(f"[OK] Excel file generated: {output_path}")
print(f"Total skills exported: {len(skills_data)}")
print(f"(37 locally installed + 1 GitHub only: Agent-Reach)")
print(f"\nNote: Added Agent-Reach from GitHub (not locally installed) - 9.4k stars!")
