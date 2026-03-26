#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""在参考 Excel 中查找 Agent-Reach"""

from openpyxl import load_workbook

# 读取参考文件
ref_path = r"C:\Users\Administrator\.openclaw\media\outbound\7edfac85-80ef-49fd-9750-05e75f1c3c3e.xlsx"

wb = load_workbook(ref_path, data_only=True)
ws = wb.active

print("查找 Agent-Reach (不区分大小写)...")
print("=" * 80)
found = False
for row in range(2, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    if name and 'agent' in str(name).lower() and 'reach' in str(name).lower():
        found = True
        print(f"\n找到于第 {row} 行:")
        for col in range(1, 12):
            col_name = ws.cell(row=1, column=col).value
            val = ws.cell(row=row, column=col).value
            print(f"  {col_name}: {val}")

if not found:
    print("\n未找到 Agent-Reach 技能")
    print("\n所有包含 'Agent' 的技能:")
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name and 'agent' in str(name).lower():
            print(f"  - {name}")
