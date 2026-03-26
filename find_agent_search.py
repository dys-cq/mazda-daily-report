#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""在参考 Excel 中查找 agent-search"""

from openpyxl import load_workbook

# 读取参考文件
ref_path = r"C:\Users\Administrator\.openclaw\media\outbound\7edfac85-80ef-49fd-9750-05e75f1c3c3e.xlsx"

wb = load_workbook(ref_path, data_only=True)
ws = wb.active

print("查找 agent-search...")
for row in range(1, ws.max_row + 1):
    first_cell = ws.cell(row=row, column=1).value
    if first_cell and 'agent-search' in str(first_cell).lower():
        print(f"\n找到于第 {row} 行:")
        for col in range(1, 12):
            col_name = ws.cell(row=1, column=col).value
            val = ws.cell(row=row, column=col).value
            print(f"  {col_name}: {val}")
