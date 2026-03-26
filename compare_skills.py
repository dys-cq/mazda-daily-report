#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""列出参考 Excel 中的所有技能名称"""

from openpyxl import load_workbook

# 读取参考文件
ref_path = r"C:\Users\Administrator\.openclaw\media\outbound\7edfac85-80ef-49fd-9750-05e75f1c3c3e.xlsx"

wb = load_workbook(ref_path, data_only=True)
ws = wb.active

print("参考文件中的所有技能名称:")
print("=" * 80)
skill_names = []
for row in range(2, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    if name:
        skill_names.append(name)
        print(f"{row-1}. {name}")

print("\n" + "=" * 80)
print(f"总计：{len(skill_names)} 个技能")

# 本地已安装的技能
local_skills = [
    "anything-to-notebooklm", "baidu-baike-data", "baidu-scholar-search-skill", "baidu-search",
    "baoyu-article-illustrator", "baoyu-comic", "baoyu-compress-image", "baoyu-cover-image",
    "baoyu-danger-gemini-web", "baoyu-danger-x-to-markdown", "baoyu-format-markdown", "baoyu-image-gen",
    "baoyu-infographic", "baoyu-markdown-to-html", "baoyu-post-to-wechat", "baoyu-post-to-wechat-backup-20260315_014801",
    "baoyu-post-to-weibo", "baoyu-post-to-x", "baoyu-slide-deck", "baoyu-translate",
    "baoyu-url-to-markdown", "baoyu-xhs-images", "baoyu-yunwu", "digest-to-email",
    "feishu-doc", "feishu-drive", "feishu-perm", "feishu-wiki",
    "html-to-media-cover", "info-designer-infographic", "markdown-mailer", "mazda-daily-report",
    "miaoda-app-builder", "multi-search-engine", "OpenClaw-DeepReeder", "skill-creator",
    "weather"
]

print("\n参考文件中有但本地没有的技能:")
print("=" * 80)
for name in skill_names:
    if name not in local_skills:
        print(f"  - {name}")
