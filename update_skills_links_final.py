#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
OpenClaw 技能管理表 - 最终更新技能链接
基于 skills.sh、GitHub、InStreet Coze、飞书开放平台等
"""

import openpyxl

# 加载现有工作簿
input_path = r"C:\Users\Administrator\.openclaw\workspace\OpenClaw 技能管理表.xlsx"
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# 最终技能链接数据（整合多个来源）
skills_links = {
    "browser-use": "https://github.com/browser-use/browser-use\nhttps://skills.sh/browser-use/browser-use/browser-use",
    "agent-browser": "https://github.com/vercel-labs/agent-browser\nhttps://skills.sh/vercel-labs/agent-browser/agent-browser",
    "active-maintenance": "https://docs.openclaw.ai\nhttps://github.com/openclaw/openclaw\nhttps://instreet.coze.site/post/6351c7a7-f9ff-4445-bdc3-1868a950cd3e",
    "agent-commons": "https://github.com/openclaw/openclaw\nhttps://docs.openclaw.ai",
    "feishu-bitable-creator": "https://open.feishu.cn/document/home/index\nhttps://open.feishu.cn/document/u-doc-api-guide\nhttps://github.com/search?q=feishu+bitable+skill",
    "ai-news-oracle": "https://clawhub.ai\nhttps://github.com/search?q=ai+news+oracle+skill",
    "bug-audit": "https://github.com/search?q=bug+audit+skill\nhttps://skills.sh/squirrelscan/skills/audit-website",
    "nodejs-project-arch": "https://github.com/search?q=nodejs+project+template\nhttps://github.com/openclaw/openclaw",
    "coze-web-search": "https://www.coze.cn\nhttps://instreet.coze.site/skills\nhttps://code.coze.cn/",
    "coze-web-fetch": "https://www.coze.cn\nhttps://instreet.coze.site/skills\nhttps://code.coze.cn/",
    "news-aggregator-skill": "https://clawhub.ai\nhttps://github.com/search?q=news+aggregator+skill",
    "feishu-sheets": "https://open.feishu.cn/document/home/index\nhttps://open.feishu.cn/document/u-sheet-api-guide\nhttps://github.com/search?q=feishu+sheets+skill",
    "frontend-design-skill": "https://github.com/anthropics/skills\nhttps://skills.sh/anthropics/skills/frontend-design",
    "agent-team-orchestration": "https://docs.openclaw.ai\nhttps://github.com/openclaw/openclaw",
    "arxiv-search-collector": "https://arxiv.org\nhttps://github.com/search?q=arxiv+search+skill",
    "三层记忆架构": "https://docs.openclaw.ai\nhttps://github.com/openclaw/openclaw\nhttps://instreet.coze.site/post/92323fa5-70ff-4399-92b1-ff0b2a7f010a",
    "OpenClaw 心跳机制": "https://docs.openclaw.ai\nhttps://github.com/openclaw/openclaw\nhttps://instreet.coze.site/post/6351c7a7-f9ff-4445-bdc3-1868a950cd3e\nhttps://instreet.coze.site/post/17353a49-af0e-4f7d-8b1a-52b5627f95e2",
    "Skill 最小可用面": "https://github.com/anthropics/skills\nhttps://skills.sh/anthropics/skills/skill-creator\nhttps://instreet.coze.site/skill.md",
    "多技能协同编排": "https://docs.openclaw.ai\nhttps://github.com/openclaw/openclaw\nhttps://instreet.coze.site/post/d3f50bc2-1865-47da-9333-a563165abc96",
    "学术雷达 V3.0": "https://clawhub.ai\nhttps://github.com/search?q=academic+radar+skill\nhttps://arxiv.org",
    "飞书文档创建最佳实践": "https://open.feishu.cn/document/home/index\nhttps://open.feishu.cn/document/u-doc-api-guide\nhttps://github.com/search?q=feishu+doc+skill"
}

# 找到技能名称列和链接列的索引
headers = [cell.value for cell in ws[1]]
skill_name_idx = headers.index("技能名称") + 1
skill_link_idx = headers.index("技能链接") + 1

# 更新链接
updated_count = 0
for row in range(2, ws.max_row + 1):
    skill_name = ws.cell(row=row, column=skill_name_idx).value
    if skill_name in skills_links:
        ws.cell(row=row, column=skill_link_idx).value = skills_links[skill_name]
        updated_count += 1

# 设置列宽和自动换行
ws.column_dimensions['G'].width = 60  # 技能链接列

from openpyxl.styles import Alignment
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=skill_link_idx)
    cell.alignment = Alignment(wrap_text=True, vertical='top')

# 保存文件
wb.save(input_path)

print(f"[OK] 技能链接最终更新完成！")
print(f"[INFO] 共更新 {updated_count} 个技能的链接")
print(f"[INFO] 文件位置：{input_path}")
print("\n[INFO] 链接来源说明:")
print("  - GitHub: https://github.com")
print("  - Skills.sh: https://skills.sh")
print("  - InStreet Coze: https://instreet.coze.site")
print("  - 飞书开放平台：https://open.feishu.cn")
print("  - arXiv: https://arxiv.org")
print("  - OpenClaw 文档：https://docs.openclaw.ai")
