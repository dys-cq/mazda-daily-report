#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""导出已安装技能列表到 Excel（带链接）"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 技能数据（包含链接）
skills = [
    {"序号": 1, "技能名称": "anything-to-notebooklm", "描述": "多源内容智能处理器：支持微信公众号、网页、YouTube、PDF、Markdown 等，自动上传到 NotebookLM 并生成播客/PPT/思维导图等多种格式", "技能链接 URL": ""},
    {"序号": 2, "技能名称": "baidu-baike-data", "描述": "百度百科查询工具，查询权威百科名词解释", "技能链接 URL": "https://clawhub.com/skills/baidu-baike-data"},
    {"序号": 3, "技能名称": "baidu-scholar-search-skill", "描述": "百度学术搜索 - 搜索中英文学术文献（期刊、会议、论文等）", "技能链接 URL": "https://clawhub.com/skills/baidu-scholar-search-skill"},
    {"序号": 4, "技能名称": "baidu-search", "描述": "使用百度 AI 搜索引擎进行网络搜索", "技能链接 URL": "https://clawhub.com/skills/baidu-search"},
    {"序号": 5, "技能名称": "baoyu-article-illustrator", "描述": "分析文章结构，识别需要配图的位置，生成类型×风格二维插图", "技能链接 URL": "https://github.com/ideacco/baoyu-skills-openclaw"},
    {"序号": 6, "技能名称": "baoyu-comic", "描述": "知识漫画创作者，支持多种艺术风格和语气，创建原创教育漫画", "技能链接 URL": "https://github.com/ideacco/baoyu-skills-openclaw"},
    {"序号": 7, "技能名称": "baoyu-compress-image", "描述": "压缩图片为 WebP（默认）或 PNG 格式，自动选择工具", "技能链接 URL": "https://github.com/ideacco/baoyu-skills-openclaw"},
    {"序号": 8, "技能名称": "baoyu-cover-image", "描述": "生成文章封面图，支持 5 维度定制（类型、调色板、渲染、文字、情绪）", "技能链接 URL": "https://github.com/ideacco/baoyu-skills-openclaw"},
    {"序号": 9, "技能名称": "baoyu-danger-gemini-web", "描述": "通过逆向 Gemini Web API 生成图像和文本", "技能链接 URL": "https://github.com/ideacco/baoyu-skills-openclaw"},
    {"序号": 10, "技能名称": "baoyu-danger-x-to-markdown", "描述": "将 X (Twitter) 推文和文章转换为 Markdown 格式", "技能链接 URL": "https://github.com/ideacco/baoyu-skills-openclaw"},
    {"序号": 11, "技能名称": "baoyu-format-markdown", "描述": "格式化纯文本或 Markdown 文件，添加 frontmatter、标题、摘要等", "技能链接 URL": "https://github.com/ideacco/baoyu-skills-openclaw"},
    {"序号": 12, "技能名称": "baoyu-image-gen", "描述": "AI 图像生成，支持 OpenAI、Google、DashScope 和 Replicate API", "技能链接 URL": "https://github.com/ideacco/baoyu-skills-openclaw"},
    {"序号": 13, "技能名称": "baoyu-infographic", "描述": "生成专业信息图，21 种布局类型和 20 种视觉风格", "技能链接 URL": "https://github.com/ideacco/baoyu-skills-openclaw"},
    {"序号": 14, "技能名称": "baoyu-markdown-to-html", "描述": "将 Markdown 转换为带样式的 HTML，支持微信兼容主题", "技能链接 URL": "https://github.com/ideacco/baoyu-skills-openclaw"},
    {"序号": 15, "技能名称": "baoyu-post-to-wechat", "描述": "微信公众号文章一键发布工具，支持专业排版、AI 封面图生成", "技能链接 URL": "https://github.com/ideacco/baoyu-skills-openclaw"},
    {"序号": 16, "技能名称": "baoyu-post-to-wechat-backup-20260315_014801", "描述": "微信公众号发布工具备份版本", "技能链接 URL": "https://github.com/ideacco/baoyu-skills-openclaw"},
    {"序号": 17, "技能名称": "baoyu-post-to-weibo", "描述": "发布内容到微博，支持普通帖子和头条文章", "技能链接 URL": "https://github.com/ideacco/baoyu-skills-openclaw"},
    {"序号": 18, "技能名称": "baoyu-post-to-x", "描述": "发布内容和文章到 X (Twitter)，支持普通帖子和 X Articles", "技能链接 URL": "https://github.com/ideacco/baoyu-skills-openclaw"},
    {"序号": 19, "技能名称": "baoyu-slide-deck", "描述": "从内容生成专业幻灯片图片", "技能链接 URL": "https://github.com/ideacco/baoyu-skills-openclaw"},
    {"序号": 20, "技能名称": "baoyu-translate", "描述": "文章和文档翻译，支持快速/普通/精翻三种模式", "技能链接 URL": ""},
    {"序号": 21, "技能名称": "baoyu-url-to-markdown", "描述": "使用 Chrome CDP 获取任何 URL 并转换为 Markdown", "技能链接 URL": "https://github.com/ideacco/baoyu-skills-openclaw"},
    {"序号": 22, "技能名称": "baoyu-xhs-images", "描述": "生成小红书信息图系列，11 种视觉风格和 8 种布局", "技能链接 URL": "https://github.com/ideacco/baoyu-skills-openclaw"},
    {"序号": 23, "技能名称": "baoyu-yunwu", "描述": "通过 Yunwu AI API 调用 Gemini Imagen 模型生成图片", "技能链接 URL": ""},
    {"序号": 24, "技能名称": "digest-to-email", "描述": "生成每日 AI/技术摘要并通过 SMTP 发送邮件", "技能链接 URL": ""},
    {"序号": 25, "技能名称": "feishu-doc", "描述": "飞书文档读写操作", "技能链接 URL": "https://github.com/autogame-17/feishu-skills"},
    {"序号": 26, "技能名称": "feishu-drive", "描述": "飞书云存储文件管理", "技能链接 URL": "https://github.com/autogame-17/feishu-skills"},
    {"序号": 27, "技能名称": "feishu-perm", "描述": "飞书文档和文件权限管理", "技能链接 URL": "https://github.com/autogame-17/feishu-skills"},
    {"序号": 28, "技能名称": "feishu-wiki", "描述": "飞书知识库导航", "技能链接 URL": "https://github.com/autogame-17/feishu-skills"},
    {"序号": 29, "技能名称": "html-to-media-cover", "描述": "HTML 转长图并适配自媒体封面尺寸", "技能链接 URL": ""},
    {"序号": 30, "技能名称": "info-designer-infographic", "描述": "高密度信息设计与视觉生成指导", "技能链接 URL": ""},
    {"序号": 31, "技能名称": "markdown-mailer", "描述": "通过 SMTP 发送 Markdown 或 HTML 内容邮件", "技能链接 URL": ""},
    {"序号": 32, "技能名称": "mazda-daily-report", "描述": "生成马自达售后日报（经营 + 客诉/线索+CSI）", "技能链接 URL": ""},
    {"序号": 33, "技能名称": "miaoda-app-builder", "描述": "在百度秒哒平台创建、修改、生成和部署网站和 Web 应用", "技能链接 URL": "https://clawhub.com/skills/miaoda-app-builder"},
    {"序号": 34, "技能名称": "multi-search-engine", "描述": "多搜索引擎集成，17 个引擎（8 个国内 +9 个全球）", "技能链接 URL": "https://clawhub.com/skills/multi-search-engine"},
    {"序号": 35, "技能名称": "OpenClaw-DeepReeder", "描述": "OpenClaw 代理的默认网页内容网关，读取 X、Reddit、YouTube 等", "技能链接 URL": "https://github.com/astonysh/OpenClaw-DeepReeder"},
    {"序号": 36, "技能名称": "skill-creator", "描述": "创建或更新 AgentSkills", "技能链接 URL": "https://clawhub.com/skills/skill-creator"},
    {"序号": 37, "技能名称": "weather", "描述": "通过 wttr.in 或 Open-Meteo 获取当前天气和预报", "技能链接 URL": "https://clawhub.com/skills/weather"},
]

# 创建 DataFrame
df = pd.DataFrame(skills)

# 导出到 Excel
output_path = r"C:\Users\Administrator\.openclaw\workspace\已安装技能列表_带链接.xlsx"

# 使用 ExcelWriter 创建文件
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name="技能列表")
    
    # 获取 workbook 和 worksheet
    wb = writer.book
    ws = wb.active
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 70
    ws.column_dimensions['D'].width = 60
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 设置数据行样式
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = data_alignment
            cell.border = border
    
    # 冻结首行
    ws.freeze_panes = 'A2'

print(f"[OK] Excel file generated: {output_path}")
print(f"Total skills exported: {len(skills)}")
