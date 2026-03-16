# -*- coding: utf-8 -*-
import os
import re
import json
import html
import shutil
import argparse
from datetime import datetime
import pandas as pd

STORES = ['重庆金团','重庆瀚达','重庆银马','重庆万事新','西藏鼎恒']
STORE_CODES = {'重庆金团':'M18003','重庆瀚达':'M23002','重庆银马':'A28856','重庆万事新':'M18571','西藏鼎恒':'M18912'}


def find_daily_dir(root='E:/'):
    # 优先使用固定路径（无空格）
    fixed = 'E:/2026年KPI/每日分析'
    if os.path.exists(fixed):
        return fixed

    base = None
    for it in os.scandir(root):
        if '2026' in it.name and 'KPI' in it.name and it.is_dir():
            base = it.path
            break
    if not base:
        raise FileNotFoundError('未找到 2026 KPI 根目录')

    daily = None
    for sub in os.scandir(base):
        if 'ÿ' in sub.name or '每' in sub.name:
            daily = sub.path
            break
    if not daily:
        raise FileNotFoundError('未找到 每日分析 目录')
    return daily


def find_latest_file(daily, keywords, exts=None):
    """查找最新文件，支持中文关键字匹配（处理 Windows 编码问题）"""
    candidates = []
    for root, _, files in os.walk(daily):
        for f in files:
            if f.startswith('~$'):
                continue
            if exts and not f.lower().endswith(tuple(exts)):
                continue
            
            try:
                fb_hex = f.encode('utf-8').hex()
                matched = False
                
                # 售后日报：检查 keywords 是否包含相关词，然后匹配 hex
                if ('售后日报' in keywords or '售后' in keywords or '日报' in keywords or '综合' in keywords):
                    if 'e594aee5908e' in fb_hex or 'e697a5e68aa5' in fb_hex or 'e7bbbce59088' in fb_hex:
                        candidates.append(os.path.join(root, f))
                        matched = True
                
                if not matched and ('CSI' in keywords or '自主调研' in keywords or '自主' in keywords):
                    if '435349' in fb_hex or 'e887aa' in fb_hex:
                        candidates.append(os.path.join(root, f))
                        matched = True
                
                if not matched and ('保险平台' in keywords or '保险' in keywords or '平台' in keywords):
                    if 'e4bf9de999a9' in fb_hex or 'e5b9b3e58fb0' in fb_hex:
                        candidates.append(os.path.join(root, f))
                        matched = True
                
                if not matched and ('客诉' in keywords or '线索' in keywords or '投诉' in keywords or '线索工单' in keywords):
                    if 'e5aea2e8af89' in fb_hex or 'e5b7a5e58d95' in fb_hex:
                        candidates.append(os.path.join(root, f))
                        matched = True
                
                # 回退：直接字符串匹配
                if not matched and any(kw in f for kw in keywords):
                    candidates.append(os.path.join(root, f))
            except Exception:
                if any(kw in f for kw in keywords):
                    candidates.append(os.path.join(root, f))
    
    if not candidates:
        return None
    return max(candidates, key=lambda p: os.path.getmtime(p))


def load_lead(lead_file):
    if not lead_file:
        return None
    try:
        tables = pd.read_html(lead_file, encoding='utf-8')
    except Exception:
        tables = pd.read_html(lead_file)
    df = max(tables, key=lambda t: t.shape[1])
    df.columns = df.iloc[0]
    df = df.iloc[1:]
    return df


def parse_lead_stats(lead_df):
    stats = {s: {'total':0,'closed':0,'unclosed':0,'timeout':0} for s in STORES}
    if lead_df is None or lead_df.empty:
        return stats

    cols = list(lead_df.columns)
    name_col = cols[15] if len(cols) > 15 else cols[-1]
    status_col = cols[8] if len(cols) > 8 else cols[0]
    timeout_col = cols[-1]

    for st in STORES:
        sub = lead_df[lead_df[name_col].astype(str).str.contains(st, na=False)]
        total = len(sub)
        closed = sub[sub[status_col].astype(str).str.contains('关闭|已|闭', na=False)].shape[0]
        unclosed = max(total - closed, 0)
        timeout = (
            pd.to_numeric(
                sub[timeout_col].astype(str).str.extract(r'(\d+\.?\d*)', expand=False),
                errors='coerce'
            ).fillna(0) > 0
        ).sum()
        stats[st] = {
            'total': int(total),
            'closed': int(closed),
            'unclosed': int(unclosed),
            'timeout': int(timeout)
        }
    return stats


def parse_chezhinet(chezhinet_df, stores):
    """解析车质网投诉数据，按门店统计"""
    stats = {s: {'total': 0, 'unclosed': 0} for s in stores}
    if chezhinet_df is None or chezhinet_df.empty:
        return stats
    
    # 车质网投诉文件结构：列 16 或 31 是经销商名称
    # 尝试多个可能的经销商列
    dealer_cols = [16, 31]
    dealer_col = None
    for col_idx in dealer_cols:
        if len(chezhinet_df.columns) > col_idx:
            sample = str(chezhinet_df.iloc[0, col_idx]) if len(chezhinet_df) > 0 else ''
            if '重庆' in sample or '西藏' in sample or '公司' in sample:
                dealer_col = col_idx
                break
    
    if dealer_col is None:
        return stats
    
    for st in stores:
        sub = chezhinet_df[chezhinet_df.iloc[:, dealer_col].astype(str).str.contains(st, na=False)]
        total = len(sub)
        # 投诉状态列可能在列 24
        status_col = 24
        if len(chezhinet_df.columns) > status_col:
            unclosed = sub[sub.iloc[:, status_col].astype(str).str.contains('未关闭|未处理', na=False)].shape[0]
        else:
            unclosed = total
        stats[st] = {'total': int(total), 'unclosed': int(unclosed)}
    
    return stats


def load_csi(csi_file):
    if not csi_file:
        return None, None
    xl = pd.ExcelFile(csi_file)
    stat = pd.read_excel(csi_file, sheet_name=0)
    bad = pd.read_excel(csi_file, sheet_name=xl.sheet_names[-1])
    return stat, bad


def parse_csi(stat_df, bad_df, csi_file=None):
    csi_stats = {}
    bad_lists = {}

    # 解析表头时间范围（CSI 统计表第一行包含表头信息，需要用 openpyxl 读取）
    csi_header_map = {}
    if csi_file and os.path.exists(csi_file):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(csi_file, data_only=True)
            ws = wb.worksheets[0]  # 第一个 sheet 是统计表
            # 读取第一行
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                if cell.value and isinstance(cell.value, str) and ('：' in cell.value or ':' in cell.value):
                    m = re.match(r'^(.*?)[：:]\s*(.*)$', cell.value)
                    if m:
                        key = m.group(1).strip()
                        value = m.group(2).strip()
                        csi_header_map[key] = value
                        print(f'  CSI Header[{col}]: {key} = {value}')
        except Exception as e:
            print(f'  Error parsing CSI header: {e}')

    for st in STORES:
        code = STORE_CODES[st]
        data = {
            '经销商名称':'',
            '评价工单结算范围':'',
            '本月维修合同数_label':'本月维修合同数',
            '本月维修合同数':0,
            '直评日期范围':'',
            '评价客户数':0,
            '参与率':'0.00%',
            '满意客户数':0,
            '满意率':'0.00%'
        }
        if stat_df is not None and not stat_df.empty and len(stat_df) > 2:
            # 第 1 行是表头，第 2 行是列名，从第 3 行开始是数据行（索引 2）
            for idx in range(2, len(stat_df)):
                row = stat_df.iloc[idx]
                # 检查是否是目标经销商（列 0 是经销商代码）
                row_code = str(row.iloc[0]) if len(row) > 0 else ''
                if code == row_code or code in row_code:
                    data['经销商名称'] = str(row.iloc[1]) if len(row) > 1 else ''
                    # 评价工单结算范围：从表头获取
                    data['评价工单结算范围'] = csi_header_map.get('评价工单结算范围', '')
                    # 本月维修合同数 + 表头时间范围
                    contract_cnt = float(row.iloc[4]) if len(row)>4 and pd.notna(row.iloc[4]) else 0
                    contract_range = csi_header_map.get('本月维修合同', '')
                    data['本月维修合同数'] = format_int(contract_cnt)
                    data['本月维修合同数_label'] = f"本月维修合同数（{contract_range}）" if contract_range else '本月维修合同数'
                    # 直评日期范围：从表头获取
                    data['直评日期范围'] = csi_header_map.get('直评日期', '')
                    eval_cnt = float(row.iloc[5]) if len(row)>5 and pd.notna(row.iloc[5]) else 0
                    part = float(row.iloc[6]) if len(row)>6 and pd.notna(row.iloc[6]) else 0
                    sat_cnt = float(row.iloc[7]) if len(row)>7 and pd.notna(row.iloc[7]) else 0
                    sat_rate = float(row.iloc[8]) if len(row)>8 and pd.notna(row.iloc[8]) else 0
                    data['评价客户数'] = format_int(eval_cnt)
                    data['参与率'] = f"{part*100:.2f}%" if part <= 1 else f"{part:.2f}%"
                    data['满意客户数'] = format_int(sat_cnt)
                    data['满意率'] = f"{sat_rate*100:.2f}%" if sat_rate <= 1 else f"{sat_rate:.2f}%"
                    break
        csi_stats[st] = data

        rows = []
        if bad_df is not None and not bad_df.empty and bad_df.shape[1] >= 4:
            sub = bad_df[bad_df.iloc[:,3].astype(str) == code]
            for _, r in sub.iterrows():
                rows.append({
                    '客诉所属经销商': str(r.iloc[6]) if len(r)>6 else '',
                    '直评时间': str(r.iloc[9]) if len(r)>9 else '',
                    '维修合同号': str(r.iloc[0]) if len(r)>0 else '',
                    '不满意点概述': str(r.iloc[13]) if len(r)>13 else ''
                })
        bad_lists[st] = rows

    return csi_stats, bad_lists


def load_biz(aftermarket_file):
    if not aftermarket_file or not os.path.exists(aftermarket_file):
        return None
    try:
        xl = pd.ExcelFile(aftermarket_file)
        target_sheet = None
        for sn in xl.sheet_names:
            if 'mazda' in str(sn).lower() or '月' in str(sn):
                target_sheet = sn
                break
        if target_sheet is None:
            target_sheet = xl.sheet_names[0]
        return pd.read_excel(aftermarket_file, sheet_name=target_sheet)
    except Exception:
        return None


def load_biz_detail(aftermarket_file, daily_dir):
    # 从综合日报第3个sheet（索引2）读取“零附件目标/达成/达成率、机油单车、事故单车、保养台次”等明细
    xlsx_files = []
    if aftermarket_file and os.path.exists(aftermarket_file):
        xlsx_files.append(aftermarket_file)
    xlsx_files += [
        os.path.join(daily_dir, f)
        for f in os.listdir(daily_dir)
        if f.lower().endswith('.xlsx') and (not f.startswith('~$'))
    ]

    for wb in xlsx_files:
        try:
            xl = pd.ExcelFile(wb)
            if len(xl.sheet_names) >= 3:
                df = pd.read_excel(wb, sheet_name=2)
                if df is not None and not df.empty and df.shape[1] >= 21:
                    return df
        except Exception:
            continue
    return None


def format_percent(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == '' or s.lower() == 'nan':
        return None

    had_percent = '%' in s
    s_num = s.replace('%', '').replace(',', '').strip()
    try:
        n = float(s_num)
    except Exception:
        return s

    # 统一两位小数：
    # - 源数据是小数（0~1）时转换成百分比
    # - 源数据已是百分值（如 86.5 或 86.5%）时直接格式化
    if (not had_percent) and n <= 1:
        n = n * 100
    return f"{n:.2f}%"


def format_int(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == '' or s.lower() == 'nan':
        return None
    try:
        n = float(str(v).replace(',', ''))
        return int(round(n))
    except Exception:
        return v


def parse_biz(biz_df, biz_detail_df=None):
    out = {}
    if biz_df is None or biz_df.empty:
        return {s:{} for s in STORES}

    # 明细sheet按“经销商简称”建索引（用于你关心的零附件/保养/机油单车/事故单车）
    detail_by_store = {}
    if biz_detail_df is not None and not biz_detail_df.empty:
        cols = list(biz_detail_df.columns)
        name_col = cols[2] if len(cols) > 2 else cols[0]
        for st in STORES:
            rr = biz_detail_df[biz_detail_df[name_col].astype(str).str.contains(st, na=False)]
            if not rr.empty:
                detail_by_store[st] = rr.iloc[0]

    for st in STORES:
        code = STORE_CODES[st]
        row = biz_df[biz_df.iloc[:,0].astype(str)==code]
        if row.empty:
            out[st] = {}
            continue

        r = row.iloc[0]
        cols = r.index

        def find_any(keys):
            for c in cols:
                if any(k in str(c) for k in keys):
                    return r[c]
            return None

        d = detail_by_store.get(st)

        def dget(idx):
            if d is None:
                return None
            try:
                return d.iloc[idx]
            except Exception:
                return None

        # 这些字段优先取综合日报“零附件明细”sheet（第3个sheet）
        month_target = dget(9)
        month_achieve = dget(10)
        month_rate = dget(11)
        quarter_target = dget(12)
        quarter_achieve = dget(13)
        quarter_rate = dget(14)
        month_maint_count = dget(16)
        oil_per_car = dget(17)
        accident_per_car = dget(20)

        # 进店台次：取自综合日报"零件 - 经销商明细"sheet 的"工单数"列（索引 21）
        # 台次达成率：从主表查找包含"台次"和"达成"的列
        repair_order_count = dget(21)  # 工单数
        visit_rate = find_any(['台次','达成'])  # 从主表查找台次达成率
        
        out[st] = {
            '服务总收入': format_int(find_any(['服务总'])),
            '零件总收入': format_int(find_any(['零件总'])),
            '工时总收入': format_int(find_any(['工时总'])),
            '进店台次（工单数）': format_int(repair_order_count if repair_order_count is not None else find_any(['进店','台','工单'])),
            '台次达成率': format_percent(visit_rate),
            '机油单车': format_int(oil_per_car if oil_per_car is not None else find_any(['机油','单车'])),
            '事故单车': format_int(accident_per_car if accident_per_car is not None else find_any(['事故','单车'])),
            '当月保养台次': format_int(month_maint_count),
            '当季度保养台次累计': format_int(month_maint_count),  # 日报仅给当月保养台次，季度累计暂按月累计口径占位
            '当月零附件目标': format_int(month_target),
            '当月零附件达成': format_int(month_achieve),
            '当月零件达成率': format_percent(month_rate),
            '当季度零附件目标': format_int(quarter_target),
            '当季度零附件达成': format_int(quarter_achieve),
            '当季度达成率': format_percent(quarter_rate),
        }
    return out


def load_platform(platform_file):
    if not platform_file:
        return None
    xls = pd.ExcelFile(platform_file)
    out = []
    for sn in xls.sheet_names:
        try:
            df = pd.read_excel(platform_file, sheet_name=sn)
            if df is not None and not df.empty and df.shape[1] >= 11:
                out.append((sn, df))
        except Exception:
            continue
    return out


def parse_platform(monthly_sheets):
    stats = {s:{} for s in STORES}
    if not monthly_sheets:
        return stats

    # 只匹配“X月”或“X月X日”命名，避免把 Sheet12 这类误判成月份sheet
    month_name_pattern = re.compile(r'^\s*\d{1,2}\s*月(\s*\d{1,2}\s*日)?\s*$', re.IGNORECASE)
    month_sheets = [(sn, df) for sn, df in monthly_sheets if month_name_pattern.match(str(sn))]
    if not month_sheets:
        # 若未命中，回退到旧逻辑（含数字）确保兼容历史命名
        month_sheets = [(sn, df) for sn, df in monthly_sheets if any(ch.isdigit() for ch in str(sn))]
    if not month_sheets:
        month_sheets = monthly_sheets

    # common layout: [0编号,1经销商代码,2经销商简称,3区域,4经理,5新保,6续保出单,7续保录单,8续保汇总,9忠诚用户,10忠诚率]
    for st in STORES:
        vals = {}
        quarter_renew_sum = 0
        monthly_renew_rates = []

        for idx_m, (sn, df) in enumerate(month_sheets):
            name_col = df.columns[2] if len(df.columns) > 2 else df.columns[0]
            row = df[df[name_col].astype(str).str.contains(st, na=False)]
            if row.empty:
                continue
            r = row.iloc[0]

            month_prefix = f"{sn}-"
            field_map = [
                (5, '新保出单'),
                (6, '续保出单'),
                (7, '续保录单'),
                (8, '续保汇总'),
                (9, '忠诚用户'),
                (10, '续保率'),
            ]
            for fidx, label in field_map:
                if len(df.columns) > fidx:
                    raw = r.iloc[fidx]
                    val = format_percent(raw) if label == '续保率' else format_int(raw)
                    vals[month_prefix + label] = val

                    if label == '续保汇总':
                        try:
                            quarter_renew_sum += int(format_int(raw) or 0)
                        except Exception:
                            pass
                    if label == '续保率':
                        rp = format_percent(raw)
                        if rp and rp != '-':
                            try:
                                monthly_renew_rates.append(float(str(rp).replace('%', '').strip()))
                            except Exception:
                                pass

            # 不输出经销商简称、区域等门店元信息，只保留统计指标

        # 当季度参考续保率 = 当季度每月继保率的平均值
        if monthly_renew_rates:
            vals['参考续保率'] = format_percent(sum(monthly_renew_rates) / len(monthly_renew_rates) / 100)
        else:
            vals['参考续保率'] = '-'

        # 也输出季度续保汇总累加，便于核对
        vals['当季度续保汇总累加'] = format_int(quarter_renew_sum)

        stats[st] = vals
    return stats


BIZ_DISPLAY_ORDER = [
    '服务总收入',
    '零件总收入',
    '工时总收入',
    '进店台次（工单数）',
    '台次达成率',
    '机油单车',
    '事故单车',
    '当月保养台次',
    '当季度保养台次累计',
    '当月零附件目标',
    '当月零附件达成',
    '当月零件达成率',
    '当季度零附件目标',
    '当季度零附件达成',
    '当季度达成率',
]


def build_markdown(run_folder, report_date, biz_metrics, platform_stats, lead_stats, csi_stats, bad_lists, chezhinet_stats=None):
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    md = [
        '# KPI 每日全维度分析报告',
        '',
        f'**报告日期**: {report_date}  ',
        f'**生成时间**: {now}  ',
        f'**数据目录**: `{run_folder}`',
        ''
    ]
    for st in STORES:
        md.append(f'## {st}')
        md.append('')
        md.append('### 经营（日常）')
        bm = biz_metrics.get(st, {})
        if bm:
            md.append('| 指标 | 数值 |')
            md.append('|---|---|')
            for k in BIZ_DISPLAY_ORDER:
                v = bm.get(k)
                if v is None or str(v) == 'nan':
                    v = '-'
                md.append(f'| {k} | {v} |')
        else:
            md.append('（无经营数据）')

        md.append('')
        md.append('### 保险平台数据')
        ps = platform_stats.get(st, {})
        if ps:
            md.append('| 指标 | 数值 |')
            md.append('|---|---|')
            for k,v in ps.items():
                md.append(f'| {k} | {v} |')
        else:
            md.append('（无保险平台数据）')

        ls = lead_stats.get(st, {'total':0,'closed':0,'unclosed':0,'timeout':0})
        md.append('')
        md.append('### 客诉/线索')
        md.append(f"- 客诉/线索总数: {ls['total']}")
        md.append(f"- 未关闭: {ls['unclosed']}")
        md.append(f"- 已关闭: {ls['closed']}")
        md.append(f"- 超时时长（处理）>0: {ls['timeout']}")

        # 车质网投诉
        md.append('')
        md.append('### 车质网投诉')
        cz = chezhinet_stats.get(st, {}) if chezhinet_stats else {}
        if cz and cz.get('total', 0) > 0:
            md.append(f"- 投诉总数：**{cz.get('total', 0)}**")
            md.append(f"- 未关闭：**{cz.get('unclosed', 0)}**")
        else:
            md.append('（无车质网未关闭客诉）')

        cs = csi_stats.get(st, {})
        md.append('')
        md.append('### CSI 自主调研')
        md.append('| 指标 | 数值 |')
        md.append('|---|---|')
        md.append(f"| 经销商名称 | {cs.get('经销商名称','')} |")
        md.append(f"| 本月维修合同数 | {cs.get('本月维修合同数_label', '本月维修合同数')}：{cs.get('本月维修合同数',0)} |")
        md.append(f"| 评价工单结算范围 | {cs.get('评价工单结算范围','')} |")
        md.append(f"| 直评日期范围 | {cs.get('直评日期范围','')} |")
        md.append(f"| 评价客户数 | {cs.get('评价客户数',0)} |")
        md.append(f"| 参与率 | {cs.get('参与率','0.00%')} |")
        md.append(f"| 满意客户数 | {cs.get('满意客户数',0)} |")
        md.append(f"| 满意率 | {cs.get('满意率','0.00%')} |")

        md.append('')
        md.append('#### 不满意客户')
        bad = bad_lists.get(st, [])
        if not bad:
            md.append('（本店无不满意客户）')
        else:
            md.append('| 客诉所属经销商 | 直评时间 | 维修合同号 | 不满意点概述 |')
            md.append('|---|---|---|---|')
            for b in bad:
                md.append(f"| {b['客诉所属经销商']} | {b['直评时间']} | {b['维修合同号']} | {b['不满意点概述']} |")
        md.append('')
        md.append('---')
        md.append('')
    return '\n'.join(md)


def build_html(report_date, now, cards_html, labels, closed, unclosed, timeout, part):
    return f"""<!doctype html>
<html><head><meta charset='utf-8'><title>KPI 每日全维度分析报告</title>
<script src='https://cdn.jsdelivr.net/npm/chart.js'></script>
<style>
body{{font-family:'Microsoft YaHei',Arial;background:#f7f8fb;padding:20px;color:#222}}
.wrap{{max-width:1200px;margin:0 auto}}
.card{{background:#fff;border-radius:12px;padding:16px;margin:14px 0;box-shadow:0 2px 10px rgba(0,0,0,.08)}}
.grid{{display:grid;grid-template-columns:1fr 1fr;gap:16px}}
@media (max-width:900px){{.grid{{grid-template-columns:1fr}}}}
table{{border-collapse:collapse;width:100%;font-size:14px}}th,td{{border:1px solid #e5e7eb;padding:8px;vertical-align:top}}th{{background:#f3f4f6;text-align:left}}
canvas{{max-height:320px}}
</style></head><body><div class='wrap'>
<h1>KPI 每日全维度分析报告</h1>
<p>报告日期：{report_date} ｜ 生成时间：{now}</p>
<section class='card'><h3>总览图表</h3><canvas id='leadChart'></canvas><br/><canvas id='csiChart'></canvas></section>
{cards_html}
</div>
<script>
new Chart(document.getElementById('leadChart'),{{type:'bar',data:{{labels:{json.dumps(labels)},datasets:[{{label:'已关闭',data:{json.dumps(closed)}}},{{label:'未关闭',data:{json.dumps(unclosed)}}},{{label:'超时>0',data:{json.dumps(timeout)}}}]}}}});
new Chart(document.getElementById('csiChart'),{{type:'line',data:{{labels:{json.dumps(labels)},datasets:[{{label:'CSI参与率(%)',data:{json.dumps(part)}}}]}}}});
</script></body></html>"""


def extract_report_date(aftermarket_file):
    """从售后日报文件名提取日期，如：售后日报 20260311.xlsx → 2026-03-11"""
    if not aftermarket_file:
        return datetime.now().strftime('%Y-%m-%d')
    fname = os.path.basename(aftermarket_file)
    # 匹配 YYYYMMDD
    m = re.search(r'(\d{4})(\d{2})(\d{2})', fname)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return datetime.now().strftime('%Y-%m-%d')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--daily-dir', default=None, help='每日分析目录，默认自动识别')
    ap.add_argument('--workspace', default='C:/Users/Administrator/.openclaw/workspace')
    args = ap.parse_args()

    daily = args.daily_dir or find_daily_dir('E:/')
    
    # 自动创建今日任务文件夹
    today = datetime.now().strftime('%Y-%m-%d')
    run_folder = os.path.join(daily, f'{today}-统计')
    os.makedirs(run_folder, exist_ok=True)
    print(f'Run folder: {run_folder}')
    print(f'Daily dir: {daily}')
    print(f'Daily dir exists: {os.path.exists(daily)}')
    print(f'Files in daily: {os.listdir(daily) if os.path.exists(daily) else "N/A"}')

    # 使用 find_latest_file 查找最新文件（不纠结文件名日期）
    print(f'\nSearching in: {daily}')
    print(f'Walking files: ')
    for r, _, fs in os.walk(daily):
        for ff in fs:
            if not ff.startswith('~$'):
                print(f'  {ff[:60]}')
    
    aftermarket_file = find_latest_file(daily, ['售后日报'], ['.xlsx', '.xls'])
    csi_file = find_latest_file(daily, ['CSI', '自主调研'], ['.xlsx', '.xls'])
    platform_file = find_latest_file(daily, ['保险平台'], ['.xlsx', '.xls'])
    lead_file = find_latest_file(daily, ['线索工单'], ['.xls', '.xlsx', '.html', '.htm'])
    # 车质网投诉：查找包含"车质"或"投诉"的文件，但不能和线索工单是同一个
    chezhinet_file = None
    for f in os.listdir(daily):
        if f.startswith('~$'): continue
        fb = f.encode('utf-8').hex()
        if ('e8bda6e8b4a8' in fb or 'e68a95e8af89' in fb) and f not in [os.path.basename(lead_file) if lead_file else '']:
            chezhinet_file = os.path.join(daily, f)
            break

    print(f'\nData files:')
    print(f'  Aftermarket: {os.path.basename(aftermarket_file) if aftermarket_file else "NOT FOUND"}')
    print(f'  CSI: {os.path.basename(csi_file) if csi_file else "NOT FOUND"}')
    print(f'  Platform: {os.path.basename(platform_file) if platform_file else "NOT FOUND"}')
    print(f'  Lead: {os.path.basename(lead_file) if lead_file else "NOT FOUND"}')
    print(f'  Chezhinet: {os.path.basename(chezhinet_file) if chezhinet_file else "NOT FOUND"}')

    # 从售后日报文件名提取报告日期
    report_date = extract_report_date(aftermarket_file)

    lead_df = load_lead(lead_file)
    if lead_df is not None:
        lead_df.to_csv(os.path.join(run_folder, 'lead_utf8_header.csv'), index=False, encoding='utf-8-sig')
    lead_stats = parse_lead_stats(lead_df)

    csi_stat, csi_bad = load_csi(csi_file)
    if csi_stat is not None:
        csi_stat.to_csv(os.path.join(run_folder, 'csi_stat.csv'), index=False, encoding='utf-8-sig')
    if csi_bad is not None:
        csi_bad.to_csv(os.path.join(run_folder, 'csi_bad.csv'), index=False, encoding='utf-8-sig')
    csi_stats, bad_lists = parse_csi(csi_stat, csi_bad, csi_file)

    # 车质网投诉统计
    chezhinet_df = None
    if chezhinet_file and chezhinet_file != lead_file:  # 避免和线索工单重复
        try:
            # 读取未关闭 sheet（通常是第 3 个 sheet，索引 2）
            if chezhinet_file.endswith('.xls'):
                chezhinet_df = pd.read_excel(chezhinet_file, sheet_name=2, engine='xlrd')
            else:
                chezhinet_df = pd.read_excel(chezhinet_file, sheet_name=2)
            print(f'  Chezhinet sheet loaded: {len(chezhinet_df)} rows')
        except Exception as e:
            print(f'  Error loading chezhinet: {e}')
    chezhinet_stats = parse_chezhinet(chezhinet_df, STORES)
    print(f'  Chezhinet stats: {chezhinet_stats}')

    platform_sheets = load_platform(platform_file)
    if platform_sheets:
        for sn, sdf in platform_sheets:
            safe_sn = ''.join(ch if ch.isalnum() else '_' for ch in str(sn))
            sdf.to_csv(os.path.join(run_folder, f'platform_{safe_sn}.csv'), index=False, encoding='utf-8-sig')
    platform_stats = parse_platform(platform_sheets)

    biz_df = load_biz(aftermarket_file)
    biz_detail_df = load_biz_detail(aftermarket_file, daily)
    biz_metrics = parse_biz(biz_df, biz_detail_df)

    md_text = build_markdown(run_folder, report_date, biz_metrics, platform_stats, lead_stats, csi_stats, bad_lists, chezhinet_stats)
    date_slug = report_date.replace('-', '')
    md_out = os.path.join(run_folder, f'KPI 日报_{date_slug}_full.md')
    with open(md_out, 'w', encoding='utf-8') as f:
        f.write(md_text)

    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    # 构建车质网 HTML 显示
    chezhinet_html = {}
    for st in STORES:
        cz = chezhinet_stats.get(st, {}) if chezhinet_stats else {}
        if cz and cz.get('total', 0) > 0:
            chezhinet_html[st] = f'投诉总数：<b>{cz.get("total", 0)}</b> | 未关闭：<b>{cz.get("unclosed", 0)}</b>'
        else:
            chezhinet_html[st] = '（无车质网未关闭客诉）'
    
    cards = []
    for st in STORES:
        bm = biz_metrics.get(st, {})
        ps = platform_stats.get(st, {})
        ls = lead_stats.get(st, {'total':0,'closed':0,'unclosed':0,'timeout':0})
        cs = csi_stats.get(st, {})
        bad = bad_lists.get(st, [])

        biz_rows = ''.join([
            f"<tr><td>{html.escape(str(k))}</td><td>{html.escape(str(bm.get(k) if (bm.get(k) is not None and str(bm.get(k)) != 'nan') else '-'))}</td></tr>"
            for k in BIZ_DISPLAY_ORDER
        ]) or '<tr><td colspan="2">（无经营数据）</td></tr>'
        platform_rows = ''.join([f"<tr><td>{html.escape(str(k))}</td><td>{html.escape(str(v))}</td></tr>" for k,v in ps.items()]) or '<tr><td colspan="2">（无保险平台数据）</td></tr>'
        csi_rows = ''.join([
            f"<tr><td>经销商名称</td><td>{html.escape(str(cs.get('经销商名称','')))}</td></tr>",
            f"<tr><td>{html.escape(cs.get('本月维修合同数_label', '本月维修合同数'))}</td><td>{html.escape(str(cs.get('本月维修合同数',0)))}</td></tr>",
            f"<tr><td>评价工单结算范围</td><td>{html.escape(str(cs.get('评价工单结算范围','')))}</td></tr>",
            f"<tr><td>直评日期范围</td><td>{html.escape(str(cs.get('直评日期范围','')))}</td></tr>",
            f"<tr><td>评价客户数</td><td>{html.escape(str(cs.get('评价客户数',0)))}</td></tr>",
            f"<tr><td>参与率</td><td>{html.escape(str(cs.get('参与率','0.00%')))}</td></tr>",
            f"<tr><td>满意客户数</td><td>{html.escape(str(cs.get('满意客户数',0)))}</td></tr>",
            f"<tr><td>满意率</td><td>{html.escape(str(cs.get('满意率','0.00%')))}</td></tr>",
        ])

        if bad:
            bad_rows = ''.join([f"<tr><td>{html.escape(b['客诉所属经销商'])}</td><td>{html.escape(b['直评时间'])}</td><td>{html.escape(b['维修合同号'])}</td><td>{html.escape(b['不满意点概述'])}</td></tr>" for b in bad])
            bad_tbl = f"<table><thead><tr><th>客诉所属经销商</th><th>直评时间</th><th>维修合同号</th><th>不满意点概述</th></tr></thead><tbody>{bad_rows}</tbody></table>"
        else:
            bad_tbl = '（本店无不满意客户）'

        cards.append(f"""
        <section class='card'>
          <h2>{st}</h2>
          <div class='grid'>
            <div>
              <h3>经营（日常）</h3>
              <table><tbody>{biz_rows}</tbody></table>
              <h3>保险平台数据</h3>
              <table><tbody>{platform_rows}</tbody></table>
            </div>
            <div>
              <h3>客诉/线索</h3>
              <ul>
                <li>客诉/线索总数：<b>{ls['total']}</b></li>
                <li>未关闭：<b>{ls['unclosed']}</b></li>
                <li>已关闭：<b>{ls['closed']}</b></li>
                <li>超时时长（处理）&gt;0：<b>{ls['timeout']}</b></li>
              </ul>
              <h3>车质网投诉</h3>
              <p>{chezhinet_html.get(st, '（无车质网未关闭客诉）')}</p>
              <h3>CSI 自主调研</h3>
              <table><tbody>{csi_rows}</tbody></table>
              <h4>不满意客户</h4>
              {bad_tbl}
            </div>
          </div>
        </section>
        """)

    labels = STORES
    closed = [lead_stats.get(s,{}).get('closed',0) for s in STORES]
    unclosed = [lead_stats.get(s,{}).get('unclosed',0) for s in STORES]
    timeout = [lead_stats.get(s,{}).get('timeout',0) for s in STORES]
    part = [float(csi_stats[s]['参与率'].strip('%')) if csi_stats.get(s) else 0 for s in STORES]

    html_text = build_html(report_date, now, ''.join(cards), labels, closed, unclosed, timeout, part)
    html_out = os.path.join(run_folder, f'KPI 日报_{date_slug}_full.html')
    with open(html_out, 'w', encoding='utf-8') as f:
        f.write(html_text)

    # mirror to workspace outputs
    ws = args.workspace
    shutil.copyfile(md_out, os.path.join(ws, f'KPI 日报_{date_slug}_full.md'))
    shutil.copyfile(html_out, os.path.join(ws, f'KPI 日报_{date_slug}_full.html'))

    print(f'\nOK: {html_out}')


if __name__ == '__main__':
    main()
